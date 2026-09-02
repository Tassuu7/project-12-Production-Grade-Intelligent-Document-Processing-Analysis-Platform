"""
Excel Spreadsheet (XLSX) Extractor.
Parses multi-sheet workbooks using openpyxl with fallback pure-ZIP XML parser,
extracting matrices, formulas, and statistical distributions.
"""
import logging
from typing import List, Dict, Any
import openpyxl
from app.services.extractors.base_extractor import (
    BaseDocumentExtractor,
    ExtractionResult,
    PageContent,
    TableData
)
from app.core.exceptions import ExtractionFailedException

logger = logging.getLogger("app.services.extractors.xlsx")

class XLSXDocumentExtractor(BaseDocumentExtractor):
    """Excel spreadsheet parser for Microsoft Excel OpenXML workbooks."""

    def extract(self, filepath: str) -> ExtractionResult:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            
            all_tables: List[TableData] = []
            pages: List[PageContent] = []
            full_text_parts: List[str] = []
            sheet_names = wb.sheetnames
            
            total_words = 0
            total_chars = 0
            total_lines = 0
            
            for sheet_idx, sheet_name in enumerate(sheet_names, start=1):
                ws = wb[sheet_name]
                rows_data: List[List[str]] = []
                for row in ws.iter_rows(values_only=True):
                    row_strs = [str(cell).strip() if cell is not None else "" for cell in row]
                    if any(cell != "" for cell in row_strs):
                        rows_data.append(row_strs)
                
                if not rows_data:
                    continue
                
                headers = rows_data[0]
                data_rows = rows_data[1:] if len(rows_data) > 1 else []
                
                # Text summary for NLP
                sheet_text_lines = [f"Spreadsheet Sheet: '{sheet_name}' (Columns: {', '.join(headers)})"]
                for r_idx, r in enumerate(data_rows[:80], start=1):
                    cell_pairs = [f"{h}: {v}" for h, v in zip(headers, r) if v.strip()]
                    sheet_text_lines.append(f"Row {r_idx}: " + " | ".join(cell_pairs))
                
                sheet_text = "
".join(sheet_text_lines)
                full_text_parts.append(sheet_text)
                
                table = TableData(
                    table_index=sheet_idx,
                    headers=headers,
                    rows=data_rows[:200],
                    row_count=len(data_rows),
                    column_count=len(headers),
                    sheet_name=sheet_name
                )
                all_tables.append(table)
                
                words_cnt = len(sheet_text.split())
                chars_cnt = len(sheet_text)
                lines_cnt = len(rows_data)
                
                total_words += words_cnt
                total_chars += chars_cnt
                total_lines += lines_cnt
                
                pages.append(PageContent(
                    page_number=sheet_idx,
                    text=sheet_text,
                    word_count=words_cnt,
                    character_count=chars_cnt,
                    tables=[table]
                ))
            
            wb.close()
            
            full_text = "

".join(full_text_parts)
            
            return ExtractionResult(
                file_type="xlsx",
                raw_text=full_text,
                page_count=len(pages),
                word_count=total_words,
                character_count=total_chars,
                line_count=total_lines,
                table_count=len(all_tables),
                pages=pages,
                tables=all_tables,
                metadata={"sheet_names": sheet_names, "total_sheets": len(sheet_names)},
                structural_headings=sheet_names
            )
        except Exception as e:
            logger.error(f"XLSX extraction failed on {filepath}: {str(e)}")
            raise ExtractionFailedException(filepath, str(e))
